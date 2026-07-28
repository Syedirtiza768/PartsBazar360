from collections import Counter
from openpyxl import load_workbook
import json
import re

source = r"C:\Users\Irtaza Hassan\Downloads\FEBEST_Availability_2026-07-14_16_09_36.xlsx"
output = r"F:\apps\PartsBazar360\tmp\catalog-architecture-analysis\artifacts\febest-oem-prefixes.json"

workbook = load_workbook(source, read_only=True, data_only=True)
sheet = workbook["Worksheet"]
prefixes = Counter()
token_count = 0

for row in sheet.iter_rows(min_row=2, min_col=6, max_col=6, values_only=True):
    value = str(row[0] or "").strip()
    for token in re.split(r"\s*,\s*", value):
        token = token.strip()
        if not token:
            continue
        token_count += 1
        words = token.split()
        first_number_word = next((i for i, word in enumerate(words) if any(c.isdigit() for c in word)), None)
        if first_number_word is None:
            prefix = token
        else:
            prefix = " ".join(words[:first_number_word])
        prefixes[prefix.upper()] += 1

payload = {
    "token_count": token_count,
    "distinct_inferred_prefixes": len(prefixes),
    "prefix_counts": [{"prefix": prefix, "count": count} for prefix, count in prefixes.most_common()],
}

with open(output, "w", encoding="utf-8") as handle:
    json.dump(payload, handle, indent=2)

print(json.dumps(payload["prefix_counts"][:120], indent=2))
