"""
Convert eBay Motors MVL workbooks (US / DE / UK / AU) → CSV for Postgres COPY.

UK files may be password-protected (try Fahrzeugliste / VehicleList).

Usage:
  python apps/api/scripts/mvl_xlsx_to_csv.py --all-downloads --out-dir tmp/mvl
  python apps/api/scripts/mvl_xlsx_to_csv.py --market UK --xlsx "..." --out tmp/uk.csv
"""
from __future__ import annotations

import argparse
import csv
import io
import re
import sys
import tempfile
import uuid
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

try:
    import msoffcrypto
except ImportError:
    msoffcrypto = None


PASSWORDS = ["VehicleList", "Fahrzeugliste", "vehiclelist", "fahrzeugliste"]


def norm(value: str | None) -> str:
    s = (value or "").upper()
    s = re.sub(r"[^A-Z0-9]+", "", s)
    return s


def cell(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s in ("N/A", "None", "null", "--", "-"):
        return ""
    return s


def unlock_if_needed(src: Path) -> Path:
    """Return a readable xlsx path (decrypt to temp if encrypted)."""
    if msoffcrypto is None:
        return src
    with open(src, "rb") as f:
        office = msoffcrypto.OfficeFile(f)
        if not office.is_encrypted():
            return src
        for pw in PASSWORDS:
            try:
                f.seek(0)
                office = msoffcrypto.OfficeFile(f)
                office.load_key(password=pw)
                bio = io.BytesIO()
                office.decrypt(bio)
                out = Path(tempfile.gettempdir()) / f"mvl_unlocked_{src.stem}.xlsx"
                out.write_bytes(bio.getvalue())
                print(f"decrypted {src.name} with password={pw} -> {out}", flush=True)
                return out
            except Exception:
                continue
        raise SystemExit(f"Could not decrypt {src} with known passwords")


def expand_years_uk(year_cell: str) -> list[int]:
    """1997|1998|1999 or single year."""
    years: list[int] = []
    for part in re.split(r"[|/,;]+", year_cell or ""):
        part = part.strip()
        if re.fullmatch(r"(?:19|20)\d{2}", part):
            years.append(int(part))
    return sorted(set(years))


def expand_years_de(period: str) -> list[int]:
    """1969/11-1973/08 or 1998/01-2015/12 → inclusive years."""
    s = (period or "").strip()
    m = re.search(r"(19|20)\d{2}.*?[-–].*?(19|20)\d{2}", s)
    if m:
        # extract first and last 4-digit years
        ys = [int(x) for x in re.findall(r"(?:19|20)\d{2}", s)]
        if len(ys) >= 2:
            a, b = ys[0], ys[-1]
            if b < a:
                a, b = b, a
            # guard absurd spans
            b = min(b, a + 60)
            return list(range(a, b + 1))
    ys = [int(x) for x in re.findall(r"(?:19|20)\d{2}", s)]
    return sorted(set(ys))


CSV_COLS = [
    "id",
    "epid",
    "kType",
    "market",
    "sourceKey",
    "year",
    "make",
    "model",
    "trim",
    "submodel",
    "engine",
    "driveType",
    "fuelType",
    "body",
    "aspiration",
    "displayName",
    "region",
    "partsModel",
    "numDoors",
    "normalizedMake",
    "normalizedModel",
    "createdAt",
    "updatedAt",
]


def write_row(w, **kwargs):
    now = kwargs.pop("now")
    make = kwargs["make"]
    model = kwargs["model"]
    n_make = norm(make)
    n_model = norm(model)
    if not n_make or not n_model:
        return False
    market = kwargs["market"]
    source_key = kwargs["sourceKey"]
    w.writerow(
        [
            str(uuid.uuid4()),
            kwargs.get("epid") or "",
            kwargs.get("kType") or "",
            market,
            source_key,
            kwargs["year"],
            make,
            model,
            kwargs.get("trim") or "",
            kwargs.get("submodel") or "",
            kwargs.get("engine") or "",
            kwargs.get("driveType") or "",
            kwargs.get("fuelType") or "",
            kwargs.get("body") or "",
            kwargs.get("aspiration") or "",
            kwargs.get("displayName") or "",
            kwargs.get("region") or "",
            kwargs.get("partsModel") or "",
            kwargs.get("numDoors") or "",
            n_make,
            n_model,
            now,
            now,
        ]
    )
    return True


def convert_us(ws, w, limit=None):
    headers = [cell(h) for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {h: i for i, h in enumerate(headers) if h}
    now = datetime.now(timezone.utc).isoformat()
    written = skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if limit is not None and written >= limit:
            break

        def g(name):
            i = idx.get(name)
            return cell(row[i]) if i is not None else ""

        epid = g("ePID")
        make = g("Make")
        model = g("Model")
        year_s = g("Year")
        try:
            year = int(year_s)
        except ValueError:
            skipped += 1
            continue
        if not epid:
            skipped += 1
            continue
        ok = write_row(
            w,
            now=now,
            market="US",
            sourceKey=f"US:{epid}",
            epid=epid,
            year=year,
            make=make,
            model=model,
            trim=g("Trim"),
            submodel=g("Submodel"),
            engine=g("Engine"),
            driveType=g("Drive Type"),
            fuelType=g("Fuel Type Name"),
            body=g("Body"),
            aspiration=g("Aspiration"),
            displayName=g("DisplayName"),
            region=g("Region"),
            partsModel=g("Parts Model"),
            numDoors=g("NumDoors") if g("NumDoors").isdigit() else "",
        )
        if ok:
            written += 1
        else:
            skipped += 1
        if written % 50000 == 0 and written:
            print(f"US written={written}", flush=True)
    return written, skipped


def convert_au(ws, w, limit=None):
    headers = [cell(h) for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {h: i for i, h in enumerate(headers) if h}
    now = datetime.now(timezone.utc).isoformat()
    written = skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if limit is not None and written >= limit:
            break

        def g(name):
            i = idx.get(name)
            return cell(row[i]) if i is not None else ""

        epid = g("ePID")
        make = g("Make")
        model = g("Model")
        year_s = g("Year")
        try:
            year = int(year_s)
        except ValueError:
            skipped += 1
            continue
        if not epid:
            skipped += 1
            continue
        ktype = g("Ktype") or g("K-Type") or g("KType")
        ok = write_row(
            w,
            now=now,
            market="AU",
            sourceKey=f"AU:{epid}",
            epid=epid,
            kType=ktype,
            year=year,
            make=make,
            model=model,
            trim=g("Variant"),
            submodel=g("Submodel"),
            engine=g("Engine"),
            body=g("Body"),
            displayName=f"{make} {model} {year}".strip(),
            partsModel=g("Plat_Gen"),
            region=g("Type"),
        )
        if ok:
            written += 1
        else:
            skipped += 1
        if written % 25000 == 0 and written:
            print(f"AU written={written}", flush=True)
    return written, skipped


def convert_uk(ws, w, limit=None):
    headers = [cell(h) for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {h: i for i, h in enumerate(headers) if h}
    now = datetime.now(timezone.utc).isoformat()
    written = skipped = 0
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if limit is not None and written >= limit:
            break

        def g(name):
            i = idx.get(name)
            return cell(row[i]) if i is not None else ""

        make = g("Make")
        model = g("Model")
        ktype = g("K-Type") or g("KType") or g("Ktype")
        years = expand_years_uk(g("Year"))
        if not make or not model or not years:
            skipped += 1
            continue
        for year in years:
            if limit is not None and written >= limit:
                break
            sk = (
                f"UK:{ktype or norm(make)+norm(model)}:{year}:"
                f"{norm(g('Type'))}:{norm(g('Variant'))}:{norm(g('Engine'))}"
            )[:220]
            if sk in seen:
                skipped += 1
                continue
            seen.add(sk)
            ok = write_row(
                w,
                now=now,
                market="UK",
                sourceKey=sk,
                kType=ktype,
                year=year,
                make=make,
                model=model,
                trim=g("Type") or g("Variant"),
                submodel=g("Variant"),
                engine=g("Engine"),
                body=g("BodyStyle"),
                displayName=f"{make} {model} {g('Type')} {year}".strip(),
                region=g("BodyStyle"),
            )
            if ok:
                written += 1
            else:
                skipped += 1
        if written and written % 50000 == 0:
            print(f"UK written={written}", flush=True)
    return written, skipped


def convert_de(ws, w, limit=None):
    headers = [cell(h) for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {h: i for i, h in enumerate(headers) if h}
    now = datetime.now(timezone.utc).isoformat()
    written = skipped = 0
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if limit is not None and written >= limit:
            break

        def g(*names):
            for name in names:
                i = idx.get(name)
                if i is not None:
                    return cell(row[i])
            return ""

        make = g("Marke_Make_EN", "Make")
        model = g("Modell_Model_EN", "Model")
        typ = g("Typ_Type_EN", "Type")
        platform = g("Plattform_Platform_EN", "Platform")
        engine = g("Motor_Engine_EN", "Engine")
        ktype = g("K-Type", "KType", "Ktype")
        years = expand_years_de(g("Baujahr_ProductionPeriod_EN", "ProductionPeriod", "Year"))
        if not make or not model or not years:
            skipped += 1
            continue
        for year in years:
            if limit is not None and written >= limit:
                break
            sk = f"DE:{ktype or norm(make)+norm(model)}:{year}:{norm(typ)}:{norm(engine)}"[:220]
            if sk in seen:
                skipped += 1
                continue
            seen.add(sk)
            ok = write_row(
                w,
                now=now,
                market="DE",
                sourceKey=sk,
                kType=ktype,
                year=year,
                make=make,
                model=model,
                trim=typ,
                submodel=platform,
                engine=engine,
                displayName=f"{make} {model} {typ} {year}".strip(),
                region=platform,
                partsModel=platform,
            )
            if ok:
                written += 1
            else:
                skipped += 1
        if written and written % 50000 == 0:
            print(f"DE written={written}", flush=True)
    return written, skipped


SHEET_BY_MARKET = {
    "US": "US_MVL_2026_05",
    "DE": "DE_MVL_2026_04",
    "UK": "UK_MVL_2026_04",
    "AU": "AU_MVL_2026_04",
}

CONVERTERS = {
    "US": convert_us,
    "DE": convert_de,
    "UK": convert_uk,
    "AU": convert_au,
}


def convert_file(market: str, xlsx: Path, out_csv: Path, limit=None):
    readable = unlock_if_needed(xlsx)
    wb = openpyxl.load_workbook(readable, read_only=True, data_only=True)
    sheet_name = SHEET_BY_MARKET[market]
    if sheet_name not in wb.sheetnames:
        # fuzzy: first sheet containing MVL
        sheet_name = next((n for n in wb.sheetnames if "MVL" in n.upper()), None)
        if not sheet_name:
            raise SystemExit(f"No MVL sheet in {xlsx}: {wb.sheetnames}")
    ws = wb[sheet_name]
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with open(out_csv, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(CSV_COLS)
        written, skipped = CONVERTERS[market](ws, w, limit=limit)
    print(f"done market={market} written={written} skipped={skipped} -> {out_csv}", flush=True)
    return written, skipped


DEFAULT_DOWNLOADS = {
    "US": Path(r"C:\Users\Irtaza Hassan\Downloads\US_MVL_2026_05.xlsx"),
    "DE": Path(r"C:\Users\Irtaza Hassan\Downloads\DE_MVL_2026_04_V2.xlsx"),
    "UK": Path(r"C:\Users\Irtaza Hassan\Downloads\UK_MVL_2026_04 (1).xlsx"),
    "AU": Path(r"C:\Users\Irtaza Hassan\Downloads\eBay-AU_Master_Vehicle_List_202604 (1) (1).xlsx"),
}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--market", choices=["US", "DE", "UK", "AU"])
    ap.add_argument("--xlsx")
    ap.add_argument("--out")
    ap.add_argument("--all-downloads", action="store_true")
    ap.add_argument("--out-dir", default="tmp/mvl")
    ap.add_argument("--limit", type=int, default=None)
    args = ap.parse_args()

    if args.all_downloads:
        out_dir = Path(args.out_dir)
        totals = {}
        for market, path in DEFAULT_DOWNLOADS.items():
            if not path.exists():
                print(f"MISSING {market} {path}", file=sys.stderr)
                continue
            totals[market] = convert_file(market, path, out_dir / f"mvl_{market}.csv", limit=args.limit)
        print("TOTALS", totals, flush=True)
        return

    if not args.market or not args.xlsx or not args.out:
        ap.error("Provide --all-downloads or --market/--xlsx/--out")
    convert_file(args.market, Path(args.xlsx), Path(args.out), limit=args.limit)


if __name__ == "__main__":
    main()
